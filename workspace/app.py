"""供应商管理系统 - Flask 主程序"""
import os
import shutil
import sqlite3
import datetime
from pathlib import Path
from werkzeug.utils import secure_filename
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from flask import (
    Flask, render_template, request, redirect, url_for,
    flash, jsonify, send_file, send_from_directory
)

app = Flask(__name__)
app.secret_key = 'supplier-mgmt-secret-key-2025'

BASE_DIR = Path(__file__).parent
DB_PATH = BASE_DIR / 'supplier.db'
UPLOAD_DIR = BASE_DIR / 'uploads'
QUOTATION_DIR = BASE_DIR / 'quotations'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp', 'bmp'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ============================================================
# 数据库工具
# ============================================================

def get_db():
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db():
    """建表 + 预置分类"""
    conn = get_db()
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            level1 TEXT NOT NULL,
            level2 TEXT NOT NULL,
            sort_order INTEGER DEFAULT 0,
            UNIQUE(level1, level2)
        );

        CREATE TABLE IF NOT EXISTS suppliers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            short_name TEXT,
            address TEXT,
            contact_person TEXT,
            contact_info TEXT,
            notes TEXT,
            created_at DATETIME DEFAULT (datetime('now','localtime')),
            updated_at DATETIME DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS supplier_categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            supplier_id INTEGER NOT NULL,
            category_id INTEGER NOT NULL,
            level3 TEXT,
            FOREIGN KEY (supplier_id) REFERENCES suppliers(id) ON DELETE CASCADE,
            FOREIGN KEY (category_id) REFERENCES categories(id) ON DELETE CASCADE,
            UNIQUE(supplier_id, category_id)
        );

        CREATE TABLE IF NOT EXISTS products (
            erp_sku TEXT PRIMARY KEY,
            supplier_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            category_id INTEGER,
            product_url TEXT,
            package_type TEXT,
            package_size TEXT,
            package_weight TEXT,
            carton_size TEXT,
            carton_quantity INTEGER,
            carton_weight TEXT,
            is_new INTEGER DEFAULT 1,
            new_product_date DATE,
            notes TEXT,
            created_at DATETIME DEFAULT (datetime('now','localtime')),
            updated_at DATETIME DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (supplier_id) REFERENCES suppliers(id) ON DELETE CASCADE,
            FOREIGN KEY (category_id) REFERENCES categories(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS price_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            erp_sku TEXT NOT NULL,
            price REAL NOT NULL,
            price_date DATE NOT NULL,
            source TEXT,
            notes TEXT,
            created_at DATETIME DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (erp_sku) REFERENCES products(erp_sku) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS product_images (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            erp_sku TEXT NOT NULL,
            filename TEXT NOT NULL,
            sort_order INTEGER DEFAULT 1,
            uploaded_at DATETIME DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (erp_sku) REFERENCES products(erp_sku) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS follow_ups (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            supplier_id INTEGER NOT NULL,
            erp_sku TEXT,
            follow_date DATE NOT NULL,
            content TEXT,
            follow_type TEXT DEFAULT '常规维护',
            is_replied INTEGER DEFAULT 0,
            next_follow_date DATE,
            created_at DATETIME DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (supplier_id) REFERENCES suppliers(id) ON DELETE CASCADE,
            FOREIGN KEY (erp_sku) REFERENCES products(erp_sku) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS quotation_files (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            supplier_id INTEGER NOT NULL,
            display_name TEXT NOT NULL,
            filename TEXT NOT NULL,
            original_name TEXT,
            notes TEXT,
            uploaded_at DATETIME DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (supplier_id) REFERENCES suppliers(id) ON DELETE CASCADE
        );
    ''')

    # ---- 预置36类目1688分类（13个优先类目在前） ----
    default_cats = [
        # == 优先类目 (sort 1-13) ==
        ('流行配饰', '发饰', 1), ('流行配饰', '帽子', 1), ('流行配饰', '围巾披肩', 1),
        ('流行配饰', '手套', 1), ('流行配饰', '腰带', 1), ('流行配饰', '眼镜太阳镜', 1),
        ('流行配饰', '丝巾', 1), ('流行配饰', '饰品配件', 1),

        ('家具', '桌椅', 2), ('家具', '床具', 2), ('家具', '收纳柜', 2),
        ('家具', '户外家具', 2), ('家具', '办公家具', 2), ('家具', '儿童家具', 2),
        ('家具', '沙发', 2), ('家具', '床垫', 2),

        ('礼品工艺品', '摆件', 3), ('礼品工艺品', '节庆礼品', 3), ('礼品工艺品', '手工DIY', 3),
        ('礼品工艺品', '水晶制品', 3), ('礼品工艺品', '工艺礼品', 3), ('礼品工艺品', '商务礼品', 3),
        ('礼品工艺品', '促销礼品', 3), ('礼品工艺品', '收藏品', 3),

        ('家居用品', '收纳用品', 4), ('家居用品', '清洁用品', 4), ('家居用品', '厨房用品', 4),
        ('家居用品', '卫浴用品', 4), ('家居用品', '雨具', 4), ('家居用品', '家纺', 4),
        ('家居用品', '家居装饰', 4), ('家居用品', '香薰蜡烛', 4),

        ('家电', '厨房小电', 5), ('家电', '生活电器', 5), ('家电', '个护电器', 5),
        ('家电', '大家电', 5), ('家电', '厨房大家电', 5), ('家电', '热水器', 5),
        ('家电', '净化除湿', 5), ('家电', '影音电器', 5),

        ('灯具照明', '商业照明', 6), ('灯具照明', 'LED灯具', 6), ('灯具照明', '室内灯具', 6),
        ('灯具照明', '室外灯具', 6), ('灯具照明', '台灯落地灯', 6), ('灯具照明', '吊灯', 6),
        ('灯具照明', '筒灯射灯', 6), ('灯具照明', '开关插座', 6),

        ('箱包', '双肩包', 7), ('箱包', '单肩包', 7), ('箱包', '拉杆箱', 7),
        ('箱包', '钱包卡包', 7), ('箱包', '化妆包', 7), ('箱包', '旅行袋', 7),
        ('箱包', '手提包', 7), ('箱包', '斜挎包', 7),

        ('办公文教', '书写工具', 8), ('办公文教', '办公纸张', 8), ('办公文教', '文件管理', 8),
        ('办公文教', '教学用品', 8), ('办公文教', '美术用品', 8), ('办公文教', '办公耗材', 8),
        ('办公文教', '计算器电子', 8), ('办公文教', '学生文具', 8),

        ('橡胶塑料', '塑料原料', 9), ('橡胶塑料', '橡胶原料', 9), ('橡胶塑料', '塑料制品', 9),
        ('橡胶塑料', '橡胶制品', 9), ('橡胶塑料', '工业用橡胶', 9), ('橡胶塑料', '塑料包装', 9),
        ('橡胶塑料', '塑料管材', 9), ('橡胶塑料', '密封件', 9),

        ('运动娱乐', '健身器材', 10), ('运动娱乐', '骑行用品', 10), ('运动娱乐', '帐篷露营', 10),
        ('运动娱乐', '垂钓用具', 10), ('运动娱乐', '户外照明', 10), ('运动娱乐', '球类运动', 10),
        ('运动娱乐', '瑜伽舞蹈', 10), ('运动娱乐', '体育用品', 10),

        ('五金工具', '手动工具', 11), ('五金工具', '电动工具', 11), ('五金工具', '锁具', 11),
        ('五金工具', '紧固件', 11), ('五金工具', '卫浴五金', 11), ('五金工具', '门窗五金', 11),
        ('五金工具', '测量工具', 11), ('五金工具', '磨具磨料', 11),

        ('玩具', '积木', 12), ('玩具', '模型', 12), ('玩具', '玩偶', 12),
        ('玩具', '拼图', 12), ('玩具', '遥控玩具', 12), ('玩具', '益智玩具', 12),
        ('玩具', '沙滩玩具', 12), ('玩具', '电动玩具', 12),

        ('汽车配件', '车用内饰', 13), ('汽车配件', '车用外饰', 13), ('汽车配件', '车载电子', 13),
        ('汽车配件', '维护工具', 13), ('汽车配件', '汽车保养品', 13), ('汽车配件', '轮胎轮毂', 13),
        ('汽车配件', '汽车灯具', 13), ('汽车配件', '汽车零部件', 13),

        # == 其余23个类目 (sort 14-36) ==
        ('农业', '粮食谷物', 14), ('农业', '种子种苗', 14), ('农业', '肥料', 14),
        ('农业', '农药', 14), ('农业', '农业机械', 14), ('农业', '园艺工具', 14),
        ('农业', '畜牧养殖', 14), ('农业', '水产', 14),

        ('服装', '女装', 15), ('服装', '男装', 15), ('服装', '童装', 15),
        ('服装', '内衣', 15), ('服装', '运动服装', 15), ('服装', '袜子手套', 15),
        ('服装', '大码服装', 15), ('服装', '民族服装', 15),

        ('美容与个人护理', '护肤品', 16), ('美容与个人护理', '彩妆', 16),
        ('美容与个人护理', '美发护发', 16), ('美容与个人护理', '口腔护理', 16),
        ('美容与个人护理', '身体护理', 16), ('美容与个人护理', '美容仪器', 16),
        ('美容与个人护理', '美甲', 16), ('美容与个人护理', '个人洗浴', 16),

        ('商务服务', '物流快递', 17), ('商务服务', '仓储服务', 17), ('商务服务', '翻译服务', 17),
        ('商务服务', '法律服务', 17), ('商务服务', '会计税务', 17), ('商务服务', '广告营销', 17),
        ('商务服务', '展会服务', 17), ('商务服务', '软件开发', 17),

        ('化工', '有机化工原料', 18), ('化工', '无机化工原料', 18), ('化工', '涂料油漆', 18),
        ('化工', '胶粘剂', 18), ('化工', '精细化学品', 18), ('化工', '添加剂', 18),
        ('化工', '表面处理', 18), ('化工', '化学试剂', 18),

        ('建筑', '装修材料', 19), ('建筑', '水泥砖瓦', 19), ('建筑', '防水材料', 19),
        ('建筑', '瓷砖地板', 19), ('建筑', '门窗', 19), ('建筑', '楼梯扶手', 19),
        ('建筑', '管材管件', 19), ('建筑', '建筑设备', 19),

        ('消费电子', '手机配件', 20), ('消费电子', '数码相机', 20), ('消费电子', '耳机音箱', 20),
        ('消费电子', '智能穿戴', 20), ('消费电子', '电脑配件', 20), ('消费电子', '移动电源', 20),
        ('消费电子', '数据线充电器', 20), ('消费电子', '存储设备', 20),

        ('电气设备', '变压器', 21), ('电气设备', '电线电缆', 21), ('电气设备', '配电柜', 21),
        ('电气设备', '开关设备', 21), ('电气设备', '电机', 21), ('电气设备', '发电机', 21),
        ('电气设备', '稳压器', 21), ('电气设备', '电气控制', 21),

        ('电子元器件', '集成电路', 22), ('电子元器件', '二三极管', 22), ('电子元器件', '电阻电容', 22),
        ('电子元器件', '传感器', 22), ('电子元器件', '连接器', 22), ('电子元器件', 'PCB板', 22),
        ('电子元器件', '继电器', 22), ('电子元器件', '晶振振荡器', 22),

        ('能源', '煤炭', 23), ('能源', '石油制品', 23), ('能源', '天然气', 23),
        ('能源', '太阳能设备', 23), ('能源', '风能设备', 23), ('能源', '生物质能', 23),
        ('能源', '电池蓄电池', 23), ('能源', '节能设备', 23),

        ('环保', '水处理设备', 24), ('环保', '空气净化', 24), ('环保', '垃圾处理', 24),
        ('环保', '环保材料', 24), ('环保', '噪声控制', 24), ('环保', '环保检测', 24),
        ('环保', '再生资源', 24), ('环保', '节能技术', 24),

        ('定制加工', '机械加工', 25), ('定制加工', '注塑加工', 25), ('定制加工', '冲压加工', 25),
        ('定制加工', '激光切割', 25), ('定制加工', '3D打印', 25), ('定制加工', '模具加工', 25),
        ('定制加工', '钣金加工', 25), ('定制加工', '表面处理加工', 25),

        ('食品饮料', '休闲零食', 26), ('食品饮料', '饮料冲调', 26), ('食品饮料', '茶叶', 26),
        ('食品饮料', '粮油调味', 26), ('食品饮料', '酒类', 26), ('食品饮料', '保健食品', 26),
        ('食品饮料', '生鲜水果', 26), ('食品饮料', '乳制品', 26),

        ('医药保健', '医疗器械', 27), ('医药保健', '保健品', 27), ('医药保健', '中药饮片', 27),
        ('医药保健', '化学药品', 27), ('医药保健', '生物制品', 27), ('医药保健', '医用耗材', 27),
        ('医药保健', '康复器材', 27), ('医药保健', '计生用品', 27),

        ('机械', '机床', 28), ('机械', '包装机械', 28), ('机械', '纺织机械', 28),
        ('机械', '食品机械', 28), ('机械', '农业机械', 28), ('机械', '工程机械', 28),
        ('机械', '塑料机械', 28), ('机械', '印刷机械', 28),

        ('矿物冶金', '钢铁材料', 29), ('矿物冶金', '有色金属', 29), ('矿物冶金', '钢材加工', 29),
        ('矿物冶金', '铁合金', 29), ('矿物冶金', '非金属矿物', 29), ('矿物冶金', '耐火材料', 29),
        ('矿物冶金', '石墨碳素', 29), ('矿物冶金', '铸造锻压', 29),

        ('包装印刷', '纸类包装', 30), ('包装印刷', '塑料包装', 30), ('包装印刷', '金属包装', 30),
        ('包装印刷', '玻璃包装', 30), ('包装印刷', '印刷服务', 30), ('包装印刷', '包装材料', 30),
        ('包装印刷', '标签标牌', 30), ('包装印刷', '胶带封箱', 30),

        ('促销品', '广告促销品', 31), ('促销品', '商务礼品', 31), ('促销品', '会议礼品', 31),
        ('促销品', '小商品', 31), ('促销品', '节庆礼品', 31), ('促销品', '定制礼品', 31),
        ('促销品', '赠品小礼品', 31), ('促销品', '宣传物料', 31),

        ('安全防护', '劳保用品', 32), ('安全防护', '安全帽', 32), ('安全防护', '防护手套', 32),
        ('安全防护', '防护服', 32), ('安全防护', '安全鞋', 32), ('安全防护', '消防设备', 32),
        ('安全防护', '安防监控', 32), ('安全防护', '交通安全', 32),

        ('服务设备', '酒店设备', 33), ('服务设备', '餐饮设备', 33), ('服务设备', '清洁设备', 33),
        ('服务设备', '制冷设备', 33), ('服务设备', '通风设备', 33), ('服务设备', '厨房设备', 33),
        ('服务设备', '商业冰柜', 33), ('服务设备', '自动售货机', 33),

        ('鞋类配件', '男鞋', 34), ('鞋类配件', '女鞋', 34), ('鞋类配件', '童鞋', 34),
        ('鞋类配件', '运动鞋', 34), ('鞋类配件', '拖鞋凉鞋', 34), ('鞋类配件', '鞋材鞋配件', 34),
        ('鞋类配件', '靴子', 34), ('鞋类配件', '鞋垫', 34),

        ('纺织皮革', '面料', 35), ('纺织皮革', '纱线', 35), ('纺织皮革', '皮革', 35),
        ('纺织皮革', '家纺面料', 35), ('纺织皮革', '里料', 35), ('纺织皮革', '辅料', 35),
        ('纺织皮革', '羽绒填充物', 35), ('纺织皮革', '纺织原料', 35),

        ('钟表珠宝', '手表', 36), ('钟表珠宝', '钟表配件', 36), ('钟表珠宝', '黄金铂金', 36),
        ('钟表珠宝', '银饰', 36), ('钟表珠宝', '钻石', 36), ('钟表珠宝', '翡翠玉器', 36),
        ('钟表珠宝', '珍珠', 36), ('钟表珠宝', '项链手链', 36),
    ]
    existing = conn.execute("SELECT COUNT(*) FROM categories").fetchone()[0]
    if existing == 0:
        conn.executemany(
            "INSERT INTO categories (level1, level2, sort_order) VALUES (?, ?, ?)",
            default_cats
        )
    conn.commit()
    conn.close()


# ============================================================
# 辅助函数
# ============================================================

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def get_latest_price(sku):
    """获取产品最新报价"""
    db = get_db()
    row = db.execute(
        "SELECT price, price_date FROM price_history WHERE erp_sku=? ORDER BY price_date DESC LIMIT 1",
        (sku,)
    ).fetchone()
    db.close()
    return row


def get_price_change(sku):
    """获取最近两次报价的变化量"""
    db = get_db()
    rows = db.execute(
        "SELECT price FROM price_history WHERE erp_sku=? ORDER BY price_date DESC LIMIT 2",
        (sku,)
    ).fetchall()
    db.close()
    if len(rows) == 2:
        return round(rows[0]['price'] - rows[1]['price'], 2)
    return None


# ============================================================
# 前端 - 自动启动页面
# ============================================================

@app.route('/')
def index():
    """仪表盘首页"""
    db = get_db()
    supplier_count = db.execute("SELECT COUNT(*) FROM suppliers").fetchone()[0]
    today = datetime.date.today()
    today_str = today.isoformat()
    month_start = today.replace(day=1).isoformat()

    # 统计卡片
    new_count = db.execute(
        "SELECT COUNT(*) FROM products WHERE is_new=1 AND new_product_date >= ?",
        (month_start,)
    ).fetchone()[0]
    pending_count = db.execute(
        "SELECT COUNT(*) FROM follow_ups WHERE is_replied=0 AND next_follow_date <= ?",
        (today_str,)
    ).fetchone()[0]

    # 逾期未跟进
    overdue = db.execute('''
        SELECT fu.*, s.name as supplier_name, p.name as product_name
        FROM follow_ups fu
        JOIN suppliers s ON fu.supplier_id = s.id
        LEFT JOIN products p ON fu.erp_sku = p.erp_sku
        WHERE fu.is_replied = 0 AND fu.next_follow_date < ?
        ORDER BY fu.next_follow_date
        LIMIT 5
    ''', (today_str,)).fetchall()

    # 本周待跟进（7天内）
    week_end = (today + datetime.timedelta(days=7)).isoformat()
    upcoming = db.execute('''
        SELECT COUNT(*) FROM follow_ups
        WHERE is_replied = 0 AND next_follow_date >= ? AND next_follow_date <= ?
    ''', (today_str, week_end)).fetchone()[0]

    # 近7天新品
    week_ago = (today - datetime.timedelta(days=7)).isoformat()
    recent_new = db.execute(
        "SELECT COUNT(*) FROM products WHERE is_new=1 AND new_product_date >= ?",
        (week_ago,)
    ).fetchone()[0]

    # 品类分布（按一级分类统计供应商数）
    cat_dist = db.execute('''
        SELECT c.level1, COUNT(DISTINCT sc.supplier_id) as cnt
        FROM supplier_categories sc
        JOIN categories c ON sc.category_id = c.id
        GROUP BY c.level1
        ORDER BY cnt DESC
        LIMIT 10
    ''').fetchall()
    max_cat = max(r['cnt'] for r in cat_dist) if cat_dist else 1

    # 最近报价变动
    recent_changes = db.execute('''
        SELECT p.erp_sku, p.name as product_name,
               s.short_name, s.name as supplier_name,
               ph.price as latest_price, ph.price_date as latest_date,
               ph2.price as prev_price
        FROM price_history ph
        JOIN products p ON ph.erp_sku = p.erp_sku
        JOIN suppliers s ON p.supplier_id = s.id
        LEFT JOIN price_history ph2 ON ph2.erp_sku = ph.erp_sku
            AND ph2.id = (SELECT MAX(id) FROM price_history WHERE erp_sku=ph.erp_sku AND id < ph.id)
        WHERE ph.id IN (SELECT MAX(id) FROM price_history GROUP BY erp_sku)
        ORDER BY ph.price_date DESC LIMIT 10
    ''').fetchall()

    db.close()
    return render_template('index.html',
                           supplier_count=supplier_count,
                           new_count=new_count,
                           pending_count=pending_count,
                           overdue=overdue,
                           upcoming=upcoming,
                           recent_new=recent_new,
                           cat_dist=cat_dist,
                           max_cat=max_cat,
                           recent_changes=recent_changes)


# ============================================================
# 供应商模块
# ============================================================

@app.route('/suppliers')
def supplier_list():
    """供应商列表"""
    search = request.args.get('search', '').strip()
    category_id = request.args.get('category_id', '').strip()
    db = get_db()

    query = '''SELECT DISTINCT s.*, GROUP_CONCAT(c.level2, '、') as cats
               FROM suppliers s
               LEFT JOIN supplier_categories sc ON s.id = sc.supplier_id
               LEFT JOIN categories c ON sc.category_id = c.id'''
    conditions = []
    params = []

    if search:
        conditions.append(
            "(s.name LIKE ? OR s.short_name LIKE ? OR s.contact_person LIKE ? OR s.contact_info LIKE ?)"
        )
        like = f'%{search}%'
        params.extend([like, like, like, like])

    if category_id:
        conditions.append("s.id IN (SELECT supplier_id FROM supplier_categories WHERE category_id=?)")
        params.append(category_id)

    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    query += " GROUP BY s.id ORDER BY s.updated_at DESC"
    suppliers = db.execute(query, params).fetchall()

    # 获取每个供应商的关联分类，转成字典列表以便模板访问
    supplier_list = []
    for s in suppliers:
        sdict = dict(s)
        cats = db.execute('''
            SELECT c.level1, c.level2, sc.level3
            FROM supplier_categories sc
            JOIN categories c ON sc.category_id = c.id
            WHERE sc.supplier_id = ?
        ''', (s['id'],)).fetchall()
        sdict['categories'] = cats
        supplier_list.append(sdict)

    categories = db.execute("SELECT * FROM categories ORDER BY sort_order, level2").fetchall()
    db.close()
    return render_template('suppliers/list.html',
                           suppliers=supplier_list,
                           categories=categories,
                           search=search,
                           selected_category=category_id)


@app.route('/suppliers/<int:sid>')
def supplier_detail(sid):
    """供应商详情"""
    db = get_db()
    supplier = db.execute("SELECT * FROM suppliers WHERE id=?", (sid,)).fetchone()
    if not supplier:
        flash("供应商不存在", "error")
        return redirect(url_for('supplier_list'))

    # 关联分类
    cats = db.execute('''
        SELECT c.*, sc.level3 FROM categories c
        JOIN supplier_categories sc ON c.id = sc.category_id
        WHERE sc.supplier_id = ?
        ORDER BY c.sort_order, c.level2
    ''', (sid,)).fetchall()

    # 关联产品
    products = db.execute('''
        SELECT p.*, c.level1, c.level2
        FROM products p
        LEFT JOIN categories c ON p.category_id = c.id
        WHERE p.supplier_id = ?
        ORDER BY p.updated_at DESC
    ''', (sid,)).fetchall()

    # 为产品附上最新报价
    products_with_price = []
    for p in products:
        pdict = dict(p)
        pdict['cat_name'] = (p['level1'] + '→' + p['level2']) if p['level1'] else None
        latest = get_latest_price(p['erp_sku'])
        pdict['latest_price'] = latest['price'] if latest else None
        pdict['latest_price_date'] = latest['price_date'] if latest else None
        products_with_price.append(pdict)

    # 跟进记录
    follow_ups = db.execute('''
        SELECT fu.*, p.name as product_name
        FROM follow_ups fu
        LEFT JOIN products p ON fu.erp_sku = p.erp_sku
        WHERE fu.supplier_id = ?
        ORDER BY fu.follow_date DESC
    ''', (sid,)).fetchall()

    # 报价表文件
    quotations = db.execute(
        "SELECT * FROM quotation_files WHERE supplier_id=? ORDER BY uploaded_at DESC", (sid,)
    ).fetchall()

    # 转成dict以避免模板中Row对象不可赋属性问题
    supplier_dict = dict(supplier)
    supplier_dict['categories'] = cats
    supplier_dict['products'] = products_with_price

    db.close()
    return render_template('suppliers/detail.html',
                           supplier=supplier_dict,
                           follow_ups=follow_ups,
                           quotations=quotations,
                           today=datetime.date.today().isoformat())


@app.route('/suppliers/new', methods=['GET', 'POST'])
@app.route('/suppliers/add', methods=['GET', 'POST'])
def supplier_add():
    """新增供应商"""
    db = get_db()
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if not name:
            flash("供应商名称不能为空", "error")
            return redirect(url_for('supplier_add'))

        short_name = request.form.get('short_name', '').strip()
        address = request.form.get('address', '').strip()
        contact_person = request.form.get('contact_person', '').strip()
        contact_info = request.form.get('contact_info', '').strip()
        notes = request.form.get('notes', '').strip()
        category_ids = request.form.getlist('category_ids')

        cur = db.execute(
            "INSERT INTO suppliers (name, short_name, address, contact_person, contact_info, notes) VALUES (?,?,?,?,?,?)",
            (name, short_name, address, contact_person, contact_info, notes)
        )
        sid = cur.lastrowid

        for cid in category_ids:
            level3 = request.form.get(f'level3_{cid}', '').strip() or None
            db.execute("INSERT INTO supplier_categories (supplier_id, category_id, level3) VALUES (?,?,?)", (sid, cid, level3))

        db.commit()
        flash("供应商添加成功", "success")
        return redirect(url_for('supplier_detail', sid=sid))

    # 按一级分类分组
    all_cats = db.execute("SELECT * FROM categories ORDER BY sort_order, level2").fetchall()
    cat_tree = {}
    for cat in all_cats:
        if cat['level1'] not in cat_tree:
            cat_tree[cat['level1']] = []
        cat_tree[cat['level1']].append(cat)
    cat_tree_list = [{'level1': k, 'categories': v} for k, v in cat_tree.items()]
    db.close()
    return render_template('suppliers/form.html',
                           supplier=None,
                           cat_tree=cat_tree_list)


@app.route('/suppliers/<int:sid>/edit', methods=['GET', 'POST'])
def supplier_edit(sid):
    """编辑供应商"""
    db = get_db()
    supplier = db.execute("SELECT * FROM suppliers WHERE id=?", (sid,)).fetchone()
    if not supplier:
        flash("供应商不存在", "error")
        return redirect(url_for('supplier_list'))

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if not name:
            flash("供应商名称不能为空", "error")
            return redirect(url_for('supplier_edit', sid=sid))

        db.execute(
            "UPDATE suppliers SET name=?, short_name=?, address=?, contact_person=?, contact_info=?, notes=?, updated_at=datetime('now','localtime') WHERE id=?",
            (name, request.form.get('short_name', '').strip(),
             request.form.get('address', '').strip(),
             request.form.get('contact_person', '').strip(),
             request.form.get('contact_info', '').strip(),
             request.form.get('notes', '').strip(), sid)
        )
        # 更新分类关联
        db.execute("DELETE FROM supplier_categories WHERE supplier_id=?", (sid,))
        for cid in request.form.getlist('category_ids'):
            level3 = request.form.get(f'level3_{cid}', '').strip() or None
            db.execute("INSERT INTO supplier_categories (supplier_id, category_id, level3) VALUES (?,?,?)", (sid, cid, level3))
        db.commit()
        flash("供应商更新成功", "success")
        return redirect(url_for('supplier_detail', sid=sid))

    # 已选分类（含level3）
    selected = db.execute("SELECT category_id, level3 FROM supplier_categories WHERE supplier_id=?", (sid,)).fetchall()
    selected_cat_ids = {r['category_id']: r['level3'] for r in selected}
    all_cats = db.execute("SELECT * FROM categories ORDER BY sort_order, level2").fetchall()
    cat_tree = {}
    for cat in all_cats:
        if cat['level1'] not in cat_tree:
            cat_tree[cat['level1']] = []
        cat_tree[cat['level1']].append(cat)
    cat_tree_list = [{'level1': k, 'categories': v} for k, v in cat_tree.items()]
    # 将 cat_ids 注入 supplier Row 以便模板访问
    supplier_dict = dict(supplier)
    supplier_dict['cat_ids'] = selected_cat_ids
    db.close()
    return render_template('suppliers/form.html',
                           supplier=supplier_dict,
                           cat_tree=cat_tree_list)


@app.route('/suppliers/<int:sid>/delete', methods=['POST'])
def supplier_delete(sid):
    db = get_db()
    db.execute("DELETE FROM suppliers WHERE id=?", (sid,))
    db.commit()
    db.close()
    flash("供应商已删除", "success")
    return redirect(url_for('supplier_list'))


# ============================================================
# 产品模块
# ============================================================

@app.route('/products')
def product_list():
    """产品列表"""
    search = request.args.get('q', '').strip() or request.form.get('search', '').strip()
    supplier_id = request.args.get('supplier_id', '').strip()
    category_id = request.args.get('category_id', '').strip()
    new_only = request.args.get('new_only', '').strip()
    sort = request.args.get('sort', '').strip()

    db = get_db()
    query = '''
        SELECT p.*, s.short_name as supplier_short, s.name as supplier_name,
               c.level1, c.level2
        FROM products p
        JOIN suppliers s ON p.supplier_id = s.id
        LEFT JOIN categories c ON p.category_id = c.id
    '''
    conditions = []
    params = []

    if search:
        conditions.append("(p.name LIKE ? OR p.erp_sku LIKE ? OR s.name LIKE ? OR s.short_name LIKE ?)")
        like = f'%{search}%'
        params.extend([like, like, like, like])

    if supplier_id:
        conditions.append("p.supplier_id = ?")
        params.append(supplier_id)

    if category_id:
        conditions.append("p.category_id = ?")
        params.append(category_id)

    if new_only:
        conditions.append("p.is_new = 1")

    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    query += " ORDER BY p.updated_at DESC"
    products = db.execute(query, params).fetchall()

    # 附上最新报价和分类名称（转dict避免Row不可赋属性）
    product_list = []
    for p in products:
        pdict = dict(p)
        pdict['cat_name'] = (p['level1'] + '→' + p['level2']) if p['level1'] else None
        latest = get_latest_price(p['erp_sku'])
        pdict['latest_price'] = latest['price'] if latest else None
        pdict['latest_price_date'] = latest['price_date'] if latest else None
        pdict['price_change'] = get_price_change(p['erp_sku'])
        product_list.append(pdict)

    # 排序
    if sort == 'price_asc':
        product_list.sort(key=lambda p: p.get('latest_price') or 999999)
    elif sort == 'price_desc':
        product_list.sort(key=lambda p: p.get('latest_price') or 0, reverse=True)
    elif sort == 'date':
        product_list.sort(key=lambda p: p.get('latest_price_date') or '', reverse=True)

    # 品类统计（只在筛选了分类时显示）
    stats = None
    if product_list:
        prices = [p.get('latest_price') for p in product_list if p.get('latest_price')]
        if prices:
            stats = {'avg': round(sum(prices)/len(prices), 2), 'min': min(prices), 'max': max(prices), 'count': len(prices)}

    suppliers = db.execute("SELECT id, short_name, name FROM suppliers ORDER BY short_name").fetchall()
    categories = db.execute("SELECT * FROM categories ORDER BY sort_order, level2").fetchall()
    db.close()
    return render_template('products/list.html',
                           products=product_list,
                           suppliers_list=suppliers,
                           categories=categories,
                           selected_supplier=supplier_id,
                           selected_category=category_id,
                           new_only=new_only,
                           sort=sort,
                           stats=stats)


@app.route('/products/<sku>')
def product_detail(sku):
    """产品详情"""
    db = get_db()
    product = db.execute('''
        SELECT p.*, s.short_name as supplier_short, s.name as supplier_name,
               s.id as supplier_id,
               c.level1, c.level2
        FROM products p
        JOIN suppliers s ON p.supplier_id = s.id
        LEFT JOIN categories c ON p.category_id = c.id
        WHERE p.erp_sku = ?
    ''', (sku,)).fetchone()
    if not product:
        flash("产品不存在", "error")
        return redirect(url_for('product_list'))

    # 报价历史
    price_history = db.execute(
        "SELECT * FROM price_history WHERE erp_sku=? ORDER BY price_date DESC",
        (sku,)
    ).fetchall()

    # 产品图片
    images = db.execute(
        "SELECT * FROM product_images WHERE erp_sku=? ORDER BY sort_order",
        (sku,)
    ).fetchall()

    # 该产品的跟进记录
    follow_ups = db.execute('''
        SELECT fu.*, s.short_name as supplier_short
        FROM follow_ups fu
        JOIN suppliers s ON fu.supplier_id = s.id
        WHERE fu.erp_sku = ?
        ORDER BY fu.follow_date DESC
    ''', (sku,)).fetchall()

    # 计算分类显示名
    product_dict = dict(product)
    product_dict['cat_name'] = (product['level1'] + '→' + product['level2']) if product['level1'] else None

    db.close()
    return render_template('products/detail.html',
                           product=product_dict,
                           price_history=price_history,
                           images=images,
                           product_follow_ups=follow_ups,
                           today=datetime.date.today().isoformat())


@app.route('/products/new', methods=['GET', 'POST'])
@app.route('/products/add', methods=['GET', 'POST'])
def product_add():
    """新增产品"""
    db = get_db()
    if request.method == 'POST':
        erp_sku = request.form.get('erp_sku', '').strip() or None
        supplier_id = request.form.get('supplier_id', '').strip()
        name = request.form.get('name', '').strip()

        if not name:
            flash("产品名称不能为空", "error")
            return redirect(url_for('product_add'))

        # 如果没有填写SKU，自动生成
        if not erp_sku:
            ts = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
            erp_sku = f"PROD-{ts}"

        # 检查SKU是否重复
        exists = db.execute("SELECT erp_sku FROM products WHERE erp_sku=?", (erp_sku,)).fetchone()
        if exists:
            flash(f"SKU '{erp_sku}' 已存在", "error")
            return redirect(url_for('product_add'))

        product_url = request.form.get('product_url', '').strip()
        category_id = request.form.get('category_id', '').strip() or None
        package_type = request.form.get('package_type', '').strip()
        package_size = request.form.get('package_size', '').strip()
        package_weight = request.form.get('package_weight', '').strip()
        carton_size = request.form.get('carton_size', '').strip()
        carton_quantity = request.form.get('carton_quantity', '').strip() or None
        carton_weight = request.form.get('carton_weight', '').strip()
        is_new = 1 if request.form.get('is_new', '1') == '1' else 0
        new_product_date = request.form.get('new_product_date', '').strip() or None
        notes = request.form.get('notes', '').strip()

        # 产品信息
        db.execute('''
            INSERT INTO products (erp_sku, supplier_id, name, category_id, product_url,
                package_type, package_size, package_weight, carton_size, carton_quantity,
                carton_weight, is_new, new_product_date, notes)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ''', (erp_sku, supplier_id, name, category_id, product_url,
              package_type, package_size, package_weight, carton_size,
              carton_quantity, carton_weight, is_new, new_product_date, notes))

        # 报价
        price = request.form.get('price', '').strip()
        if price:
            price_date = request.form.get('price_date', '').strip() or datetime.date.today().isoformat()
            price_source = request.form.get('price_source', '').strip()
            price_notes = request.form.get('price_notes', '').strip()
            db.execute(
                "INSERT INTO price_history (erp_sku, price, price_date, source, notes) VALUES (?,?,?,?,?)",
                (erp_sku, float(price), price_date, price_source, price_notes)
            )

        # 图片上传
        files = request.files.getlist('images')
        for idx, f in enumerate(files):
            if f and f.filename and allowed_file(f.filename):
                filename = secure_filename(f.filename)
                sku_dir = UPLOAD_DIR / erp_sku
                sku_dir.mkdir(parents=True, exist_ok=True)
                f.save(str(sku_dir / filename))
                db.execute(
                    "INSERT INTO product_images (erp_sku, filename, sort_order) VALUES (?,?,?)",
                    (erp_sku, filename, idx + 1)
                )

        db.commit()
        flash("产品添加成功", "success")
        return redirect(url_for('product_detail', sku=erp_sku))

    suppliers = db.execute("SELECT id, short_name, name FROM suppliers ORDER BY short_name").fetchall()
    categories = db.execute("SELECT * FROM categories ORDER BY sort_order, level2").fetchall()
    db.close()
    return render_template('products/form.html',
                           product=None,
                           suppliers=suppliers,
                           categories=categories,
                           selected_supplier=request.args.get('supplier_id', ''))


@app.route('/products/<sku>/edit', methods=['GET', 'POST'])
def product_edit(sku):
    """编辑产品"""
    db = get_db()
    product = db.execute("SELECT * FROM products WHERE erp_sku=?", (sku,)).fetchone()
    if not product:
        flash("产品不存在", "error")
        return redirect(url_for('product_list'))

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if not name:
            flash("产品名称不能为空", "error")
            return redirect(url_for('product_edit', sku=sku))

        db.execute('''
            UPDATE products SET supplier_id=?, name=?, category_id=?, product_url=?,
                package_type=?, package_size=?, package_weight=?, carton_size=?,
                carton_quantity=?, carton_weight=?, is_new=?, new_product_date=?,
                notes=?, updated_at=datetime('now','localtime')
            WHERE erp_sku=?
        ''', (
            request.form.get('supplier_id', '').strip(),
            name,
            request.form.get('category_id', '').strip() or None,
            request.form.get('product_url', '').strip(),
            request.form.get('package_type', '').strip(),
            request.form.get('package_size', '').strip(),
            request.form.get('package_weight', '').strip(),
            request.form.get('carton_size', '').strip(),
            request.form.get('carton_quantity', '').strip() or None,
            request.form.get('carton_weight', '').strip(),
            1 if request.form.get('is_new', '1') == '1' else 0,
            request.form.get('new_product_date', '').strip() or None,
            request.form.get('notes', '').strip(),
            sku
        ))

        # 新报价
        price = request.form.get('price', '').strip()
        if price:
            price_date = request.form.get('price_date', '').strip() or datetime.date.today().isoformat()
            price_source = request.form.get('price_source', '').strip()
            price_notes = request.form.get('price_notes', '').strip()
            db.execute(
                "INSERT INTO price_history (erp_sku, price, price_date, source, notes) VALUES (?,?,?,?,?)",
                (sku, float(price), price_date, price_source, price_notes)
            )

        # 新图片
        files = request.files.getlist('images')
        if files and any(f and f.filename for f in files):
            max_order = db.execute(
                "SELECT COALESCE(MAX(sort_order),0) FROM product_images WHERE erp_sku=?", (sku,)
            ).fetchone()[0]
            for idx, f in enumerate(files):
                if f and f.filename and allowed_file(f.filename):
                    filename = secure_filename(f.filename)
                    sku_dir = UPLOAD_DIR / sku
                    sku_dir.mkdir(parents=True, exist_ok=True)
                    f.save(str(sku_dir / filename))
                    db.execute(
                        "INSERT INTO product_images (erp_sku, filename, sort_order) VALUES (?,?,?)",
                        (sku, filename, max_order + idx + 1)
                    )

        db.commit()
        flash("产品更新成功", "success")
        return redirect(url_for('product_detail', sku=sku))

    suppliers = db.execute("SELECT id, short_name, name FROM suppliers ORDER BY short_name").fetchall()
    categories = db.execute("SELECT * FROM categories ORDER BY sort_order, level2").fetchall()
    db.close()
    return render_template('products/form.html',
                           product=product,
                           suppliers=suppliers,
                           categories=categories,
                           selected_supplier='')


@app.route('/products/<sku>/delete', methods=['POST'])
def product_delete(sku):
    """删除产品"""
    db = get_db()
    # 删除图片文件夹
    sku_dir = UPLOAD_DIR / sku
    if sku_dir.exists():
        shutil.rmtree(str(sku_dir))
    db.execute("DELETE FROM products WHERE erp_sku=?", (sku,))
    db.commit()
    db.close()
    flash("产品已删除", "success")
    return redirect(url_for('product_list'))


# ============================================================
# Excel 导入
# ============================================================

EXCEL_HEADERS = [
    'ERP-SKU', '产品名称', '供应商名称', '分类（一级→二级）', '1688链接',
    '单套包装方式', '单套包装尺寸', '单套重量', '外箱尺寸', '每箱套数',
    '外箱重量', '报价（元）', '报价日期', '新品', '新品上架日期', '备注'
]


@app.route('/products/template')
def excel_template():
    """下载Excel模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "产品导入模板"

    # 表头样式
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = None  # openpyxl borders if wanted

    for col_idx, h in enumerate(EXCEL_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 示例数据
    example = [
        'TOY-2401-001', '消防局积木套装', 'XX玩具有限公司', '玩具→积木',
        'https://detail.1688.com/...', '彩盒', '35×25×8', '580g',
        '72×52×42', '24', '14.5kg', '25.80', '2025-01-15',
        '是', '2025-01-10', '满1000可议价'
    ]
    for col_idx, val in enumerate(example, 1):
        ws.cell(row=2, column=col_idx, value=val).alignment = Alignment(vertical='center')

    # 列宽
    widths = [16, 20, 18, 18, 30, 12, 12, 10, 12, 10, 10, 10, 12, 8, 14, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='产品导入模板.xlsx')


@app.route('/products/import', methods=['POST'])
def excel_import():
    """导入Excel产品数据"""
    file = request.files.get('file')
    if not file or not file.filename.endswith('.xlsx'):
        flash("请上传 .xlsx 文件", "error")
        return redirect(url_for('product_list'))

    try:
        wb = load_workbook(filename=BytesIO(file.read()), data_only=True)
    except Exception:
        flash("Excel 文件损坏或格式错误", "error")
        return redirect(url_for('product_list'))

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # 跳过表头

    if not rows:
        flash("Excel 文件没有数据", "error")
        wb.close()
        return redirect(url_for('product_list'))

    db = get_db()
    # 预加载供应商映射 {name: id} + {short_name: id}
    suppliers = db.execute("SELECT id, name, short_name FROM suppliers").fetchall()
    name_to_id = {}
    short_to_id = {}
    for s in suppliers:
        name_to_id[s['name']] = s['id']
        if s['short_name']:
            short_to_id[s['short_name']] = s['id']

    # 预加载分类映射 {"玩具→积木": category_id}
    cats = db.execute("SELECT id, level1, level2 FROM categories").fetchall()
    cat_map = {f"{c['level1']}→{c['level2']}": c['id'] for c in cats}

    success = 0
    failures = []

    for row_idx, row in enumerate(rows):
        line = row_idx + 2  # Excel行号
        vals = [str(c).strip() if c else '' for c in row]
        if len(vals) < 16:
            vals.extend([''] * (16 - len(vals)))

        sku, name, supplier_name, cat_str, url, pkg_type, pkg_size, pkg_weight, \
            carton_size, carton_qty, carton_weight, price_str, price_date, is_new_str, new_date, notes = vals

        # 必填检查
        if not name:
            failures.append(f"第{line}行：产品名称为空")
            continue
        if not supplier_name:
            failures.append(f"第{line}行：供应商名称为空")
            continue

        # 供应商匹配
        sid = name_to_id.get(supplier_name) or short_to_id.get(supplier_name)
        if not sid:
            failures.append(f"第{line}行：供应商「{supplier_name}」不存在")
            continue

        # SKU处理
        if not sku:
            sku = f"PROD-{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}-{line}"
        existing = db.execute("SELECT erp_sku FROM products WHERE erp_sku=?", (sku,)).fetchone()
        if existing:
            failures.append(f"第{line}行：SKU「{sku}」已存在")
            continue

        # 分类匹配
        cat_id = cat_map.get(cat_str) if cat_str else None

        # is_new
        is_new = 1 if is_new_str and is_new_str in ('是', '1', 'yes', 'Yes', 'YES') else 0

        # 价格
        price = float(price_str) if price_str else None
        price_date = price_date if price_date else datetime.date.today().isoformat()

        # 插入产品
        db.execute('''
            INSERT INTO products (erp_sku, supplier_id, name, category_id, product_url,
                package_type, package_size, package_weight, carton_size, carton_quantity,
                carton_weight, is_new, new_product_date, notes)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ''', (sku, sid, name, cat_id, url or None,
              pkg_type or None, pkg_size or None, pkg_weight or None,
              carton_size or None, carton_qty or None if carton_qty else None,
              carton_weight or None, is_new, new_date or None, notes or None))

        if price:
            db.execute(
                "INSERT INTO price_history (erp_sku, price, price_date) VALUES (?,?,?)",
                (sku, price, price_date)
            )

        success += 1

    db.commit()
    db.close()

    if success > 0:
        flash(f"✅ 导入成功 {success} 条" + (f"，失败 {len(failures)} 条" if failures else ""), "success")
    if failures:
        for fmsg in failures[:10]:  # 最多显示前10条失败信息
            flash(fmsg, "error")
        if len(failures) > 10:
            flash(f"... 还有 {len(failures) - 10} 条失败", "error")

    return redirect(url_for('product_list'))


@app.route('/products/<sku>/delete-image/<int:img_id>', methods=['POST'])
def image_delete(sku, img_id):
    """删除产品图片"""
    db = get_db()
    img = db.execute("SELECT * FROM product_images WHERE id=? AND erp_sku=?", (img_id, sku)).fetchone()
    if img:
        filepath = UPLOAD_DIR / sku / img['filename']
        if filepath.exists():
            filepath.unlink()
        db.execute("DELETE FROM product_images WHERE id=?", (img_id,))
        db.commit()
    db.close()
    flash("图片已删除", "success")
    return redirect(url_for('product_detail', sku=sku))


@app.route('/products/<sku>/upload-image', methods=['POST'])
def product_upload_image(sku):
    """为产品上传图片"""
    db = get_db()
    product = db.execute("SELECT erp_sku FROM products WHERE erp_sku=?", (sku,)).fetchone()
    if not product:
        db.close()
        flash("产品不存在", "error")
        return redirect(url_for('product_list'))

    files = request.files.getlist('images')
    if not files or not any(f and f.filename for f in files):
        flash("请选择图片", "error")
        db.close()
        return redirect(url_for('product_detail', sku=sku))

    max_order = db.execute(
        "SELECT COALESCE(MAX(sort_order),0) FROM product_images WHERE erp_sku=?", (sku,)
    ).fetchone()[0]

    for idx, f in enumerate(files):
        if f and f.filename and allowed_file(f.filename):
            filename = secure_filename(f.filename)
            sku_dir = UPLOAD_DIR / sku
            sku_dir.mkdir(parents=True, exist_ok=True)
            f.save(str(sku_dir / filename))
            db.execute(
                "INSERT INTO product_images (erp_sku, filename, sort_order) VALUES (?,?,?)",
                (sku, filename, max_order + idx + 1)
            )

    db.commit()
    db.close()
    flash("图片上传成功", "success")
    return redirect(url_for('product_detail', sku=sku))


@app.route('/uploads/<path:filepath>')
def uploaded_file(filepath):
    """提供图片访问"""
    return send_from_directory(str(UPLOAD_DIR), filepath)


# ============================================================
# 产品对比
# ============================================================

@app.route('/products/compare')
def product_compare():
    """产品对比"""
    ids_param = request.args.get('ids', '')
    skus = [s.strip() for s in ids_param.split(',') if s.strip()]
    if len(skus) < 2:
        flash("请至少选择 2 个产品进行对比", "error")
        return redirect(url_for('product_list'))
    if len(skus) > 5:
        skus = skus[:5]
        flash("最多对比 5 个产品", "info")

    db = get_db()
    products = []
    for sku in skus:
        p = db.execute('''
            SELECT p.*, s.name as supplier_name,
                   c.level1, c.level2
            FROM products p
            JOIN suppliers s ON p.supplier_id = s.id
            LEFT JOIN categories c ON p.category_id = c.id
            WHERE p.erp_sku = ?
        ''', (sku,)).fetchone()
        if p:
            pdict = dict(p)
            pdict['cat_name'] = (p['level1'] + '→' + p['level2']) if p['level1'] else None
            latest = get_latest_price(sku)
            pdict['latest_price'] = latest['price'] if latest else None
            pdict['latest_price_date'] = latest['price_date'] if latest else None
            products.append(pdict)

    if len(products) < 2:
        flash("所选产品不足 2 个", "error")
        db.close()
        return redirect(url_for('product_list'))

    # 找出最低价
    min_price = min((p.get('latest_price') or float('inf')) for p in products)

    db.close()
    return render_template('products/compare.html',
                           products=products,
                           min_price=min_price)


# ============================================================
# 报价模块
# ============================================================

@app.route('/products/<sku>/add-price', methods=['GET', 'POST'])
def price_add(sku):
    """新增报价记录"""
    db = get_db()
    product = db.execute("SELECT p.*, s.name as supplier_name, c.level1, c.level2 FROM products p JOIN suppliers s ON p.supplier_id=s.id LEFT JOIN categories c ON p.category_id=c.id WHERE p.erp_sku=?", (sku,)).fetchone()
    if not product:
        db.close()
        flash("产品不存在", "error")
        return redirect(url_for('product_list'))

    # 获取最新报价
    latest = db.execute("SELECT price FROM price_history WHERE erp_sku=? ORDER BY price_date DESC LIMIT 1", (sku,)).fetchone()
    latest_price = latest['price'] if latest else None

    if request.method == 'POST':
        price = request.form.get('price', '').strip()
        if not price:
            flash("报价不能为空", "error")
            db.close()
            return redirect(url_for('product_detail', sku=sku))

        price_date = request.form.get('price_date', '').strip() or datetime.date.today().isoformat()
        source = request.form.get('source', '').strip()
        notes = request.form.get('notes', '').strip()

        db.execute(
            "INSERT INTO price_history (erp_sku, price, price_date, source, notes) VALUES (?,?,?,?,?)",
            (sku, float(price), price_date, source, notes)
        )
        db.commit()
        db.close()
        flash("报价已添加", "success")
        return redirect(url_for('product_detail', sku=sku))

    db.close()
    return render_template('products/add_price.html', product=product, latest_price=latest_price, today=datetime.date.today().isoformat())


@app.route('/products/<sku>/delete-price/<int:pid>', methods=['POST'])
def price_delete(sku, pid):
    """删除报价记录"""
    db = get_db()
    db.execute("DELETE FROM price_history WHERE id=?", (pid,))
    db.commit()
    db.close()
    flash("报价记录已删除", "success")
    return redirect(url_for('product_detail', sku=sku))


# ============================================================
# 跟进记录模块
# ============================================================

@app.route('/follow-ups/new', methods=['GET', 'POST'])
def follow_up_add_form():
    """新增跟进记录"""
    db = get_db()
    supplier_id = request.args.get('supplier_id', '')

    if request.method == 'POST':
        sid = request.form.get('supplier_id', '').strip()
        if not sid:
            flash("请选择供应商", "error")
            db.close()
            return redirect(url_for('supplier_list'))
        sid = int(sid)

        follow_date = request.form.get('follow_date', '').strip() or datetime.date.today().isoformat()
        content = request.form.get('content', '').strip()
        follow_type = request.form.get('follow_type', '').strip() or '常规维护'
        erp_sku = request.form.get('erp_sku', '').strip() or None
        is_replied = 1 if request.form.get('is_replied') == '1' else 0
        next_follow_date = request.form.get('next_follow_date', '').strip() or None

        db.execute(
            "INSERT INTO follow_ups (supplier_id, erp_sku, follow_date, content, follow_type, is_replied, next_follow_date) VALUES (?,?,?,?,?,?,?)",
            (sid, erp_sku, follow_date, content, follow_type, is_replied, next_follow_date)
        )
        db.commit()
        db.close()
        flash("跟进记录已添加", "success")
        return redirect(url_for('supplier_detail', sid=sid))

    # GET: render form
    supplier = None
    if supplier_id:
        supplier = db.execute("SELECT id, name FROM suppliers WHERE id=?", (supplier_id,)).fetchone()
    products = db.execute("SELECT erp_sku, name FROM products ORDER BY name").fetchall()
    suppliers_list = db.execute("SELECT id, name FROM suppliers ORDER BY name").fetchall()
    db.close()
    return render_template('follow_ups/form.html',
                           supplier=supplier,
                           products=products,
                           suppliers_list=suppliers_list,
                           today=datetime.date.today().isoformat())


@app.route('/follow-ups/<int:fid>/delete', methods=['POST'])
def follow_up_delete(fid):
    """删除跟进记录"""
    db = get_db()
    fu = db.execute("SELECT supplier_id FROM follow_ups WHERE id=?", (fid,)).fetchone()
    if fu:
        sid = fu['supplier_id']
        db.execute("DELETE FROM follow_ups WHERE id=?", (fid,))
        db.commit()
        flash("跟进记录已删除", "success")
    db.close()
    return redirect(url_for('supplier_detail', sid=sid))


# ============================================================
# 报价表文件管理
# ============================================================

@app.route('/suppliers/<int:sid>/quotation/upload', methods=['POST'])
def quotation_upload(sid):
    """上传供应商报价表"""
    file = request.files.get('file')
    if not file or not file.filename:
        flash("请选择文件", "error")
        return redirect(url_for('supplier_detail', sid=sid))

    display_name = request.form.get('display_name', '').strip()
    if not display_name:
        flash("请输入报价表名称", "error")
        return redirect(url_for('supplier_detail', sid=sid))

    notes = request.form.get('notes', '').strip()
    today_str = datetime.date.today().isoformat()
    ext = file.filename.rsplit('.', 1)[-1] if '.' in file.filename else 'dat'
    
    # 获取供应商名称用于文件命名
    db = get_db()
    s = db.execute("SELECT name FROM suppliers WHERE id=?", (sid,)).fetchone()
    supplier_name = s['name'] if s else f'Supplier{sid}'
    # 只过滤文件名非法字符，保留中文
    def safe_name(s):
        for ch in r'\/:*?"<>|':
            s = s.replace(ch, '_')
        return s.strip()
    safe_sname = safe_name(supplier_name)
    safe_dname = safe_name(display_name)
    filename = f"{safe_sname}_{safe_dname}_{today_str}.{ext}"
    filename = safe_name(filename)  # 最终确保无非法字符

    # 保存文件到平层目录
    QUOTATION_DIR.mkdir(parents=True, exist_ok=True)
    file.save(str(QUOTATION_DIR / filename))

    # 写数据库
    db.execute(
        "INSERT INTO quotation_files (supplier_id, display_name, filename, original_name, notes) VALUES (?,?,?,?,?)",
        (sid, display_name, filename, file.filename, notes)
    )
    db.commit()
    db.close()
    flash("报价表上传成功", "success")
    return redirect(url_for('supplier_detail', sid=sid))


@app.route('/quotations/<path:filepath>')
def quotation_serve(filepath):
    """打开报价表文件（用系统默认程序）"""
    import os as _os
    full_path = QUOTATION_DIR / filepath
    if full_path.exists():
        _os.startfile(str(full_path))
    return redirect(request.referrer or url_for('supplier_list'))


@app.route('/quotations/<int:qid>/delete', methods=['POST'])
def quotation_delete(qid):
    """删除报价表文件"""
    db = get_db()
    qf = db.execute("SELECT * FROM quotation_files WHERE id=?", (qid,)).fetchone()
    if qf:
        sid = qf['supplier_id']
        filepath = QUOTATION_DIR / qf['filename']
        if filepath.exists():
            filepath.unlink()
        db.execute("DELETE FROM quotation_files WHERE id=?", (qid,))
        db.commit()
        flash("报价表已删除", "success")
    db.close()
    return redirect(url_for('supplier_detail', sid=sid))


@app.route('/quotations/open-folder')
def quotation_open_folder():
    """打开报价表文件夹"""
    import subprocess
    QUOTATION_DIR.mkdir(parents=True, exist_ok=True)
    subprocess.Popen(['explorer', str(QUOTATION_DIR)])
    flash("已打开报价表文件夹", "success")
    return redirect(request.referrer or url_for('supplier_list'))


# ============================================================
# 新品动态台
# ============================================================

@app.route('/new-products')
def new_products():
    """新品动态台"""
    month_filter = request.args.get('month', '').strip()
    db = get_db()

    query = '''
        SELECT p.*, s.short_name, s.name as supplier_name,
               c.level1, c.level2
        FROM products p
        JOIN suppliers s ON p.supplier_id = s.id
        LEFT JOIN categories c ON p.category_id = c.id
        WHERE p.is_new = 1
    '''
    params = []
    if month_filter:
        query += " AND p.new_product_date LIKE ?"
        params.append(f'{month_filter}%')

    query += " ORDER BY p.new_product_date DESC, s.short_name"
    products = db.execute(query, params).fetchall()

    # 按供应商分组
    grouped = {}
    for p in products:
        pdict = dict(p)
        latest = get_latest_price(p['erp_sku'])
        pdict['latest_price'] = latest['price'] if latest else None
        pdict['latest_price_date'] = latest['price_date'] if latest else None
        # 检查是否有跟进记录
        fu = db.execute(
            "SELECT COUNT(*) as cnt FROM follow_ups WHERE erp_sku=? ORDER BY follow_date DESC LIMIT 1",
            (p['erp_sku'],)
        ).fetchone()
        pdict['follow_status'] = '已跟进' if (fu and fu['cnt'] > 0) else '待跟进'
        key = (p['supplier_id'], p['supplier_name'], p['short_name'])
        if key not in grouped:
            grouped[key] = []
        grouped[key].append(pdict)

    # 无新品的供应商，检查其跟进状态
    all_suppliers = db.execute("SELECT * FROM suppliers ORDER BY short_name").fetchall()
    no_new_suppliers = []
    for s in all_suppliers:
        has_new = any(k[0] == s['id'] for k in grouped.keys())
        if not has_new:
            last_fu = db.execute(
                "SELECT * FROM follow_ups WHERE supplier_id=? ORDER BY follow_date DESC LIMIT 1",
                (s['id'],)
            ).fetchone()
            no_new_suppliers.append({
                'supplier': s,
                'last_follow_up': last_fu
            })

    # 组装成模板期望的 supplier_blocks 结构
    supplier_blocks = []
    for (sid, sname, sshort), products_list in grouped.items():
        # 获取该供应商的跟进提醒
        upcoming = db.execute(
            "SELECT fu.*, p.name as product_name FROM follow_ups fu LEFT JOIN products p ON fu.erp_sku=p.erp_sku WHERE fu.supplier_id=? AND fu.next_follow_date IS NOT NULL AND fu.next_follow_date >= ? ORDER BY fu.next_follow_date",
            (sid, datetime.date.today().isoformat())
        ).fetchall()
        supplier_blocks.append({
            'supplier': {'id': sid, 'name': sname, 'short_name': sshort},
            'new_products': products_list,
            'upcoming_follow_ups': upcoming,
            'latest_follow': upcoming[0] if upcoming else None
        })

    # 没有新品的供应商
    for item in no_new_suppliers:
        s = item['supplier']
        lfu = item['last_follow_up']
        supplier_blocks.append({
            'supplier': {'id': s['id'], 'name': s['name'], 'short_name': s['short_name']},
            'new_products': [],
            'upcoming_follow_ups': [],
            'latest_follow': lfu
        })

    today = datetime.date.today().isoformat()
    db.close()
    return render_template('new_products.html',
                           supplier_blocks=supplier_blocks,
                           today=today,
                           month_filter=month_filter)


# ============================================================
# 搜索 (JSON API)
# ============================================================

@app.route('/api/categories')
def api_categories():
    """级联筛选API：返回一级类目列表，或指定一级下的二级列表"""
    level1 = request.args.get('level1', '').strip()
    db = get_db()
    if level1:
        cats = db.execute(
            "SELECT id, level2 FROM categories WHERE level1=? ORDER BY sort_order, level2",
            (level1,)
        ).fetchall()
        result = [{'id': c['id'], 'level2': c['level2']} for c in cats]
    else:
        cats = db.execute(
            "SELECT DISTINCT level1, MIN(sort_order) as so FROM categories GROUP BY level1 ORDER BY so"
        ).fetchall()
        result = [{'level1': c['level1']} for c in cats]
    db.close()
    return jsonify(result)


# ============================================================
# 搜索 (JSON API)
# ============================================================

@app.route('/api/search')
def api_search():
    """统一搜索API"""
    q = request.args.get('q', '').strip()
    if len(q) < 1:
        return jsonify({'products': [], 'suppliers': []})

    db = get_db()
    like = f'%{q}%'
    products = db.execute('''
        SELECT p.erp_sku, p.name, s.short_name
        FROM products p JOIN suppliers s ON p.supplier_id = s.id
        WHERE p.name LIKE ? OR p.erp_sku LIKE ?
        LIMIT 10
    ''', (like, like)).fetchall()

    suppliers = db.execute('''
        SELECT id, name, short_name FROM suppliers
        WHERE name LIKE ? OR short_name LIKE ? OR contact_person LIKE ?
        LIMIT 10
    ''', (like, like, like)).fetchall()
    db.close()

    return jsonify({
        'products': [dict(p) for p in products],
        'suppliers': [dict(s) for s in suppliers]
    })


# ============================================================
# 启动
# ============================================================

if __name__ == '__main__':
    init_db()
    print("启动服务器: http://localhost:5000")
    app.run(debug=True, host='127.0.0.1', port=5000)
